# -*- coding: utf-8 -*-

# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html
import openpyxl


class FangtianxiaPipeline(object):

    def process_item(self, item, spider):
        writer = openpyxl.load_workbook(r"F:\PyCharmProjects\FangTianXia\FangTianXia\data_.xlsx")
        table = writer["data_info"]
        table = writer.active
        nrows = table.max_row
        ncolumns = table.max_column
        def foo(key, value, number):
            return {
                key: lambda key: table.cell(row=nrows + 1, column=number, value=value),
            }[key](number)

        for index in range(1, ncolumns + 1):
            number = [i for i, x in enumerate(item.keys()) if x == table.cell(row=2, column=index).value][0]
            foo(key=list(item.keys())[number], value=list(item.values())[number], number=index)
            writer.save(r"F:\PyCharmProjects\FangTianXia\FangTianXia\data_.xlsx")
        return item
